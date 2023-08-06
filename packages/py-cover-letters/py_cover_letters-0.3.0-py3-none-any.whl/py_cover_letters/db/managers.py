import uuid
from pathlib import Path
from typing import Protocol, List, Optional, Dict, Union

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from py_cover_letters.db.models import CoverLetter
from py_cover_letters.constants import COLUMN_MAPPING, SHEET_NAME, DEFAULT_DB_BACKUP_FOLDER
from py_cover_letters.enums import FilterType
from py_cover_letters.exceptions import CoverLetterException, UnsupportedOperationException
from py_cover_letters.utils import backup_file


class CoverLetterManager(Protocol):

    def get(self, cover_letter_id: int) -> CoverLetter:
        ...

    def create(self, cover_letter: CoverLetter) -> CoverLetter:
        ...

    def delete(self, cover_letter: CoverLetter) -> bool:
        ...

    def update(self, project: CoverLetter) -> CoverLetter:
        ...

    def list(self) -> List[CoverLetter]:
        ...

    def filter(self, filter_type: FilterType) -> List[CoverLetter]:
        ...


class ExcelManager:
    def __init__(self, filename: Path, column_mapping: Optional[Dict[int, str]] = None,
                 sheet_name: str = SHEET_NAME, backup_folder: Optional[Path] = None):
        if backup_folder is None:
            self.backup_folder = DEFAULT_DB_BACKUP_FOLDER
        else:
            self.backup_folder = backup_folder
        self.filename = filename
        if column_mapping is None:
            self.column_mapping = COLUMN_MAPPING
        else:
            self.column_mapping = column_mapping
        self.sheet_name = sheet_name
        self.columns = [col_name for _, col_name in self.column_mapping.items()]
        self.cover_letters: List[CoverLetter] = list()
        if not self.filename.exists():
            self.write_template()
        else:
            self.cover_letters = self.load()

    def write_template(self):
        if self.filename.exists():
            error_msg = f'Cannot overwrite template {self.filename}'
            raise CoverLetterException(error_msg)

        wb = Workbook()
        sheet = wb.create_sheet(self.sheet_name, 0)
        row = 1
        col = 1
        for column_name in self.columns:
            sheet.cell(row=row, column=col, value=column_name)
            col += 1
        row += 1

        wb.save(self.filename)

    def load(self) -> List[CoverLetter]:
        cover_letters = list()
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        last_row = sheet.max_row + 1
        for row in range(2, last_row):
            cover_letter_dict = dict()
            for col, name in self.column_mapping.items():
                cell_obj = sheet.cell(row=row, column=col)
                value = cell_obj.value
                cover_letter_dict[name] = value
            try:
                cover_letter = CoverLetter(**cover_letter_dict)
                cover_letters.append(cover_letter)
            except Exception as e:
                error_message = f'Unexpected error on row {row}. Type: {e.__class__.__name__} Error: {e}'
                raise CoverLetterException(error_message)
        return cover_letters

    def filter(self, filter_type: FilterType) -> List[CoverLetter]:
        filter_result = list()
        if filter_type == FilterType.COVER_LETTER_NOT_CREATED:
            filter_result = [x for x in self.cover_letters if x.date_generated is None]
        else:
            error_message = f'Unsupported FilterType {filter_type}.'
            raise UnsupportedOperationException(error_message)
        return filter_result

    def set_unique_ids(self, commit=True) -> List[CoverLetter]:
        not_saved = [x for x in self.cover_letters if x.id is None]
        for cover_letter in not_saved:
            cover_letter.id = uuid.uuid4().int
            # print(cover_letter)
        if commit:
            backup_file(self.filename, self.backup_folder)
            self.filename.unlink(missing_ok=True)
            self.write_template()
            self.save()
            # self.cover_letters = self.load()

        return not_saved

    def get(self, cover_letter_id: int) -> Union[CoverLetter, None]:
        matching = [x for x in self.cover_letters if x.id == 0]
        if len(matching) > 1:
            error_message = f'More than one cover letter has the same id. ' \
                            f'Id {cover_letter_id} count ({len(matching)})'
            raise CoverLetterException(error_message)
        elif len(matching) == 1:
            return matching[0]
        return None

    def save(self) -> None:
        wb: Workbook = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        row = sheet.max_row + 1
        for cover_letter in self.cover_letters:
            for col, attribute_name in self.column_mapping.items():
                value = getattr(cover_letter, attribute_name)
                sheet.cell(row=row, column=col, value=value)
            row += 1
        wb.save(self.filename)
