from pathlib import Path
from typing import Dict, List, Any, Optional

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .models import CoverLetter
from ..constants import COLUMN_MAPPING
from ..exceptions import CoverLetterException


class ExcelCoverLetterManager:

    def __init__(self, filename: Path, column_mapping: Optional[Dict[int, str]] = None,
                 sheet_name: str = 'Cover letters'):
        self.filename = filename
        if column_mapping is None:
            self.column_mapping = COLUMN_MAPPING
        else:
            self.column_mapping = column_mapping
        self.columns = [col_name for _, col_name in self.column_mapping.items()]
        self.sheet_name = sheet_name

    def write_template(self):
        if self.filename.exists():
            error_msg = f'Cannot overwrite template {self.filename}'
            raise CoverLetterException(error_msg)

        wb = Workbook()
        sheet = wb.create_sheet(self.sheet_name, 0)
        row = 1
        col = 1
        for column_name in self.columns:
            sheet.cell(row=row, column=col, value=column_name)
            col += 1
        row += 1

        wb.save(self.filename)

    def read(self) -> List[CoverLetter]:
        cover_letters = list()
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        last_row = sheet.max_row + 1
        for row in range(2, last_row):
            cover_letter_dict = dict()
            for col, name in self.column_mapping.items():
                cell_obj = sheet.cell(row=row, column=col)
                value = cell_obj.value
                cover_letter_dict[name] = value
            try:
                cover_letter = CoverLetter(**cover_letter_dict)
                cover_letters.append(cover_letter)
            except Exception as e:
                error_message = f'Unexpected error on row {row}. Type: {e.__class__.__name__} Error: {e}'
                raise CoverLetterException(error_message)
        return cover_letters

    def add(self, cover_letters: List[CoverLetter]):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        row = sheet.max_row + 1
        for cover_letter in cover_letters:
            for col, attribute_name in self.column_mapping.items():
                value = getattr(cover_letter, attribute_name)
                sheet.cell(row=row, column=col, value=value)
            row += 1
        wb.save(self.filename)


def excel_to_list(filename: Path, sheet_name: str, column_mapping: Dict[str, Any]) -> List[Any]:
    cover_letters = list()
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    last_row = sheet.max_row + 1
    for row in range(2, last_row):
        cover_letter_dict = dict()
        for col, name in column_mapping.items():
            cell_obj = sheet.cell(row=row, column=col)
            value = cell_obj.value
            cover_letter_dict[name] = value
        cover_letters.append(cover_letter_dict)
    return cover_letters
