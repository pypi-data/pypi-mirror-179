"""firebehaviour.py

Defines the classes used to analyse fire behaviour with PyroPy.

"""
from __future__ import annotations

from dataclasses import dataclass
import warnings
from openpyxl import load_workbook
from pandas import DataFrame, read_csv
from copy import deepcopy
from datetime import datetime

if __name__ == '__main__':
    import spreadmodels as fbs
    import helpers as fbh
else:
    from . import spreadmodels as fbs
    from . import helpers as fbh

#TODO copmare fbcalc params to other params
class Incident(object):
    """A wildfire incident.

    Attributes:
        df (Dataframe): Weather and model output data
        waf (float): wind reduction factor (0-6)
        fuel_load (float): fine fuel load (t/ha)
        fhs_surf (float): surface fuels hazard score (1-4)
        fhs_n_surf (float): near surface fuels hazard score (1-4)
        fuel_height_ns (float): near surface fuel height (cm)
        fuel_height_u (float): understorey fuel height (m) TODO calc this
        fhr_surf: Surface Fuel Hazard Rating (L, M, H, V, E)
        fhr_n_surf: Near Surface Fuel Hazard Rating (L, M, H, V, E)

    """
    def __init__(self, weather_df: DataFrame):
        self.df = weather_df
        self.waf = None
        self.fuel_load = None #t/ha
        self.fhs_surf = None
        self.fhs_n_surf = None 
        self.fuel_height_ns = None
        self.fuel_height_u = None
        self.fhr_surf = None
        self.fhr_n_surf = None
        self.grass_state = None
        self.curing = None
        self.cover_o = None
        self.height_o = None

        #TODO put these in a yaml file
        self.fbcalc_params_mk5 = {
            'waf': 'J4',
            'fuel_load': 'C4', 
        }

        self.fbcalc_params_vesta = {
            'fhs_surf': 'D6',
            'fhs_n_surf': 'D7',
            'fuel_height_ns': 'C8',
        }

        self.fbcalc_params_vesta2 = {
            'waf': 'O10',
            'fuel_load': 'F7',
            'fuel_height_u': 'C10',
        }

        # self.fbcalc_params_grass = {
        #     'grass_state': '???', #This won't work for the way FBCalc is configured
        #     'curing': 'F7',
        # }

        self.fbcalc_params_mallee = {
            'cover_o': 'E4',
            'height_o': 'E5',
        }

    def copy(self) -> Incident:
        """Deep copy.

        Returns:
            Incident: a deep copy of the `Incident`.
        """
        return deepcopy(self)

    def trim_by_datetime(self, 
            start: str, 
            end: str,
            format = '%Y%m%d %H:%M'
        ) -> None:
        """Trims the records of the `Incident Dataframe` to a datetime range.

        Args:
            start (str): the start datetime
            end (str): the end datetime
            format (str, optional): The datetime format. Defaults to '%Y%m%d %H:%M'.
        """
        start = datetime.strptime(start, format)
        end = datetime.strptime(end, format)
        self.df =  fbh.trim_by_datetimes(self.df, start, end)
        pass

    def thin_by_timestep(self, time_step: float=1) -> None:
        """Removes records from the incident dataframe.

        The dataframe will have a a minimum time step given by `time_step`

        Args:
            time_step (float, optional): Minimum time step in hours. Defaults to 1.
        """
        self.df = fbh.thin_by_timestep(self.df, time_step)

    def adjust_precision(self, precision_dict: dict) -> None:
        """Adjusts the precision of the records in the incident dataframe.

        Args:
            precision_dict (dict): defines the precision of the fields to be 
                changed. Example:
                    `precision_dict = {
                        0: ['ffdi','fros', 'ros'],
                        1: ['mc', 'mf'],
                        2: []
                    }`
        """
        self.df = fbh.adjust_precision(self.df, precision_dict)
    
    def get_params(self) -> dict:
        """Gets the model parameters that have been defined.

        Returns:
            dict: dictionary with model parameter names and values
        """
        params = list(self.__dict__.items())
        params = dict(params[1:]) #drop the dataframe
        params = {
            key: val for key, val in params.items() 
            if isinstance(val,(int,float,str))
        }
        return params

    def get_df(self) -> DataFrame:
        """

        Returns:
            DataFrame: the Incident Data as a pandas `Dataframe`
        """
        return self.df

    def get_models(self) -> list:
        """

        Returns:
            list: a list of the models that have been run for the Incident
        """
        models = {
            'forest_mk5': 'fros_mk5',
            'forest_vesta': 'fros_vesta',
        }

        return [
            key for key, val in models.items() 
            if val in self.df.columns.values
        ]

    def set_params(self, params: dict) -> None:
        """Update several model parameters using a dictionary.

        The dictionary keys must match the name of the parameter.

        Args:
            params (dict): a dictionary of the model parameters to be updated.
        """
        for key, val in params.items():
            setattr(self, key, val)

    def run_forest_mk5(self) -> None:
        """Runs the McArthur Mk5 Forest Fire Danger Meter model.

        Requires that the Wind Adjustment Factor `'waf'` and Fine fuel load 
        (t/ha) `'fuel_load'` parameters are set. See `Incident.set_params()`

        Adds the results to the `Incident.df`
        """
        forest_mk5_params = {
            'waf': self.waf,
            'fuel_load': self.fuel_load,
        }
        if self.check_params(forest_mk5_params):
            self.df = fbs.ros_forest_mk5(self.df, self.waf, self.fuel_load)

    def run_forest_vesta(self, version_12 = True) -> None:
        """Runs the Project Vesta (fuel hazard scores) model.

        Uses the Fuel Hazard Scores (FHS) described in:

        Hines, F. et al. 2010,  Overall Fuel Hazard Assessment Guide. 
        Dept of Natural Resources and Environment, East Melbourne, Vic.

        Requires that the surface FHS `'fhs_surf'`, the near
        surface FHS `'fhs_n_surf'` and the near surface fuel 
        height (cm) `'fuel_height_ns'` parameters are set. 
        See `Incident.set_params()`

        Adds the results to the `Incident.df`

        Args:
            version_12: if `True` uses the Cheney et al. 2012 equation, 
                if `False` uses the Gould et al. 2008 version. 
                Defaults to `True`.           
        """
        forest_vesta_params = {
            'fhs_surf': self.fhs_surf,
            'fhs_n_surf': self.fhs_n_surf,
            'fuel_height_ns': self.fuel_height_ns,
        }


        if self.check_params(forest_vesta_params):
            self.df = fbs.ros_forest_vesta(
                self.df, 
                self.fhs_surf, 
                self.fhs_n_surf, 
                self.fuel_height_ns,
                version_12,
            )


    def run_forest_vesta_fhr(self) -> None:
        """Runs the Project Vesta (fuel hazard ratings) model.
    
        Uses the Fuel Hazard Ratings (FHR) described in:

        Hines, F. et al. 2010,  Overall Fuel Hazard Assessment Guide. 
        Dept of Natural Resources and Environment, East Melbourne, Vic.

        Requires that the surface FHR `'fhr_surf'` and the near
        surface FHR `'fhr_n_surf'` parameters are set. 
        See `Incident.set_params()`

        Adds the results to the `Incident.df`
        """
        forest_vesta_params = {
            'fhr_surf': self.fhr_surf,
            'fhr_n_surf': self.fhr_n_surf,
        }

        if self.check_params(forest_vesta_params):
            self.df = fbs.ros_forest_vesta_fhr(
                self.df, 
                self.fhr_surf, 
                self.fhr_n_surf, 
            )

    def run_forest_vesta2(self) -> None:
        """Runs the Project Vesta Mk2 model.

        Cruz, et al. 2022 'An Empirical-Based Model for Predicting the Forward 
        Spread Rate of Wildfires in Eucalypt Forests'. 
        International Journal of Wildland Fire. https://doi.org/10.1071/WF21068

        Requires that the Wind Adjustment Factor `'waf'`, surface fuel load 
        (t/ha) `'fuel_load'` and the understorey fuel height `'fuel_height_u'` 
        parameters are set. See `Incident.set_params()`

        Adds the results to the `Incident.df`
        """
        forest_vesta2_params = {
            'waf': self.waf,
            'fuel_load': self.fuel_load,
            'fuel_height_u': self.fuel_height_u,
        }


        if self.check_params(forest_vesta2_params):
            self.df = fbs.ros_forest_vesta2(
                self.df, 
                self.waf, 
                self.fuel_load, 
                self.fuel_height_u,
            )

    def run_grass(self) -> None:
        """Runs the Cheney et al. 1998 grass and woodland model.

        Requires that the grass fuel state (N, G, E, W or F) `'grass_state'` 
        and curing (%) `'curing'` parameters are set. 
        See `Incident.set_params()`

        Adds the results to the `Incident.df`
        """
        grass_params = {
            'grass_state': self.grass_state,
            'curing': self.curing,
        }

        if self.check_params(grass_params):
            self.df = fbs.ros_grass(
                self.df,
                self.grass_state,
                self.curing,
            )

    def run_mallee(self) -> None:
        """Runs the Cruz et al 2013 Semi-arid mallee heath model.

        Requires that the overstorey cover(%) `'os_cover'` 
        and overstorey height (m) parameters are set. 
        See `Incident.set_params()`

        Adds the results to the `Incident.df`
        """
        mallee_params = {
            'cover_o': self.cover_o,
            'height_o': self.height_o,
        }

        if self.check_params(mallee_params):
            self.df = fbs.ros_mallee(
                self.df,
                self.cover_o,
                self.height_o,
            )

    def get_spread_direction(self) -> None:
        """Add the fire spread direction measured in degrees to the incident df.
        """
        self.df['spread_dir'] = fbs.spread_direction(self.df)


    def print(self, head=False) -> None:
        """Prints the field headings and rows of the `Dataframe`

        Args:
            head (bool, optional): Print only the head (first 5 rows).
                Defaults to False.
        """
        if head: print(self.df.head())
        else: print(self.df)

    def check_params(self, params: dict) -> bool:
        """Checks to see if parameters have been defined.

        Args:
            params (dict): a dictionary with the parmeters to check.

        Returns:
            bool: `True` is values for the parameters have been defined,
                else `False`
        """
        incident_params = self.get_params()
        for key in params.keys():
            if not key in incident_params.keys():
                warnings.warn(f'{key} not set - run set params')
                return False
        return True

    def compare_fbcalc(self, fn: str, models: list) -> dict:
        """Loads results from an FireBehaviourCalcs spreadsheet into the 
        `Incident.df`.

        Only loads the pages from FireBehaviourCalcs that are included in 
        `models` then adds fros values to `Incident.df` in the form 
        `'fros_{model}_fbcalc'`

        Also creates a dictionary of the fbcalc parameters relevant to the models 

        Args:
            fn (str): path to the FireBehaviourCalcs spreadsheet
            model (list): a list of the models to compare. Valid items
                include: `'mk5'` McArthur Mk5, `'vesta'` Project 
                VESTA, `'vesta2'` Vesta MkII (dry), `'mallee'`
                the mallee-heath. 
                
                Also there is a little easter egg here and if you 
                put `'mc_v'` as the model it will get the moisture content (%)
                from the VESTA model, while `'mc_v2'` will get the moisture
                content from the Vesta II model and `'mc_m'` will get the 
                moisture content from the Mallee-Heath model.

        Returns:
            dict: dictionary of fbcalc model params
        """
        fbcalc_refs = {
            'mk5': ['Forest(McArthur)', 'O', self.fbcalc_params_mk5],
            'vesta': ['Forest(VESTA)', 'P', self.fbcalc_params_vesta],
            'vesta2': ['VESTA Mk2 Dry', 'Y', self.fbcalc_params_vesta2],
            'mallee': ['Mallee-Heath', 'T', self.fbcalc_params_mallee],
            'mc_v': ['Forest(VESTA)', 'M', {}],
            'mc_v2': ['VESTA Mk2 Dry', 'N', {}],
            'mc_m': ['Mallee-Heath', 'L', {}],
        }

        self.fbcalc_params = {}

        # TODO this need to happen after the params are set.
        # run_model_functions = {
        #     'mk5': self.run_forest_mk5,
        #     'vesta': self.run_forest_vesta,
        #     'vesta2': self.run_forest_vesta2,
        # }

        if fbh.check_filepath(fn, suffix='xlsm'):
            wb = load_workbook(fn, data_only=True) #, keep_vba=True
            for model, ref in fbcalc_refs.items():
                if model in models:
                    sheet_name, column, model_params = ref
                    ws = wb[sheet_name]
                    column = ws[column]
                    ros_vals =  [cell.value for cell in column if isinstance(
                            cell.value, (float, int)
                        )]
                    field = f'fros_{model}_fbcalc'
                    if not model_params: field = f'{model}_fbcalc' #not a fros model
                    self.df[field] = ros_vals
                    if model_params: self.df[field] = self.df[field].astype(int)

                    
                    params = {param: ws[address].value 
                        for (param, address) in model_params.items()
                    }

                    for key, value in params.items():
                        self.fbcalc_params[key] = value

                    # self.set_params(params)
                    # if model in run_model_functions.keys(): #trick it into getting columns that aren't models like mc
                    #     run_model_functions[model]()            
        return self.fbcalc_params

    def compare_amicus(self, fn: str, model: str) -> None:
        if fbh.check_filepath(fn, suffix='csv'):
            df = read_csv(fn, header=0, encoding=fbh.check_encoding(fn))
            self.df['mc_amicus'] = df['Predicted FMC (%)']
            self.df[f'fros_{model}_amicus'] = df['Rate of spread (m/h)']
        pass

    def set_fbcalc(self, fn: str) -> bool:
        """Writes the weather data and model parameters to a FireBehaviourCalc
        macro enabled spreadsheet.

        ** WARNING! THE VALUES IN THE SPREADSHEET WILL BE OVERWRITTEN **

        Spreadsheet macros will not run (ie changes will not take place in 
        the spreadsheet) until you open it and click on the relevant sheets.

        Args:
            fn (str): path to the FireBehaviourCalc spreadsheet. 
                 
        
        Returns:
            `True` f successful else `False`
        """

        models = {
            'forest_mk5': ['Forest(McArthur)', self.fbcalc_params_mk5],
            'forest_vesta': ['Forest(VESTA)', self.fbcalc_params_vesta],
        }

        if fbh.check_filepath(fn, suffix='xlsm'):
            datetime_format: str = "%d/%m/%Y %H:%M"
            # weather = self.df.iloc[:, 0:6]
            weather = self.df.copy(deep=True)
            weather['date'] = weather['date_time'].dt.strftime(datetime_format.split()[0])
            weather['time'] = weather['date_time'].dt.strftime(datetime_format.split()[1])
            # re-order and remove unwanted cols
            weather_cols = ['date', 'time', 'temp', 'humidity', 'wind_dir', 'wind_speed', 'drought']
            weather = weather[weather_cols]

            wb = load_workbook(fn, read_only=False, keep_vba=True)

            #load weather data to fbcalc
            ws = wb['Weather_Site']
            row_ctr = 0
            ws_weather_cols = [2,8]
            ws_weather_rows = [12, 12+weather.shape[0]-1]
            for row in ws.iter_rows(
                min_col=ws_weather_cols[0],
                max_col=ws_weather_cols[1],
                min_row=ws_weather_rows[0],
                max_row=ws_weather_rows[1],
            ):
                cell_ctr = 0
                for cell in row:
                    cell.value = weather.iloc[row_ctr, cell_ctr]
                    cell_ctr += 1
                row_ctr += 1

            #load parameters to fbcalc
            for model in self.get_models():
                ws = wb[models[model][0]]
                for param, addr in models[model][1].items():
                    ws[addr] = self.__dict__[param]

            return True

        return False            



if __name__ == '__main__':
    pass