"""weatherdata.py

Functions for reading, transforming and writing weather data.

The fire behaviour models require certain fields to be present in the weather 
data and these are mapped using dictionaries. There are several standard 
mappings included (see below), but for non-standard data users must create 
their own dictionary.

For use in `spreadmodels` functions, the weather dataframe must include the 
following fields (column headings) as a minimum:
```
    'date_time': a datetime field
    'temp': Air temerature (°C)
    'humidity': Relative humidity (%)
    'wind_speed': 10 m wind speed (km/h)
    'wind_dir': Wind direction (°)
```
The datetime field may be replaced by two separate fields:
```
    'date': the local date
    'time': the local time
```
Other optional fields used by the spreadmodels include:
```
    'drought': Drought Factor
    'ffdi': McArthur Mk5 Forest Fire Danger Index
    'gfdi': McArthur Mk5 Grass Fire Danger Index
```

There are functions to handle the standard data formats such as BOM gridded 
weather with mapping dictionaries already defined.

```python
gridded_to_df(fn)
amicus_to_df(fn)
df_to_gridded(fn)
df_to_amicus(fn)
```

For example, if you are using BOM gridded weather data the you should run:
```python
df = gridded_to_df('your_file_path.csv')
```
"""

import pandas as pd
from pandas import DataFrame
from pathlib import Path
from openpyxl import load_workbook

if __name__ == '__main__':
    from helpers import check_filepath, check_encoding
else:
    from .helpers import check_filepath, check_encoding

FIELDS_BASE = {
    'date_time': 'Date time',
    'temp': 'Air temperature (C)',
    'humidity': 'Relative humidity (%)',
    'wind_speed': '10 m wind speed (km/h)',
    'wind_dir': 'Wind direction',
}

FIELDS_GRIDDED = {
    'date': 'Local Date',
    'time': 'Local Time',
    'temp': 'Temp (C)',
    'humidity': 'RH (%)',
    'wind_dir': 'Wind Dir',
    'wind_speed': 'Wind Speed (km/h)',
    'drought': 'Drought Factor',
    'ffdi': 'FFDI',
    'gfdi': 'GFDI',
}

FIELDS_AMICUS = {
    'date_time': 'Date time',
    'temp': 'Air temperature (°C)',
    'humidity': 'Relative humidity (%)',
    'wind_speed': '10 m wind speed (km/h)',
    'wind_dir': 'Wind direction (°)' #TODO replace all degree symbols
}

def weather_to_df(
        fn: str,
        header: int = 0, 
        col_names: dict = FIELDS_BASE, 
        datetime_format: str = "%d/%m/%Y %H:%M",
    ) -> DataFrame:

    """Reads weather forecast or observations into a pandas `DataFrame`.
    Unless a `col_names` dictionary is supplied, the `FIELDS_BASE` dictionary is 
    used.
    ```
    FIELDS_BASE = {
        'date_time': 'Date time',
        'temp': 'Air temperature (C)',
        'humidity': 'Relative humidity (%)',
        'wind_speed': '10 m wind speed (km/h)',
        'wind_dir': 'Wind direction',
    }
    ```

    Args:
        fn: the path to the csv or axf file
        header: the line containing the column headers
        col_names: the dictionary containing the column names to read into the
            dataframe
        datetime_format: string format for the date times. Use a single string 
            of the form `'date_format time_format'` even if supplying date and 
            time as two separete fields.

    Returns:
        weather dataframe
    """

    check_filepath(fn, suffix='csv')
    #swap the keys and vals in the dictionary
    col_names = {y:x for x,y in col_names.items()}
    
    # avoid UnicodeDecodeError:
    df = pd.read_csv(fn, header=header, encoding=check_encoding(fn))
    
    df = df.rename(columns=col_names)

    #just want a single datetime column containing datetime objects
    if set(['date','time']).issubset(df.columns):
        df['date_time'] = df[["date", "time"]].agg(" ".join, axis = 1)
        df = df.drop(['date','time'], axis='columns')
    if 'date_time' in df.columns:
        if datetime_format:
            df['date_time'] = pd.to_datetime(df['date_time'],format=datetime_format)
        else:
            df['date_time'] = pd.to_datetime(df['date_time'],infer_datetime_format=True)
    
    #TODO if wind dir is str then change to degrees
    # if 'wind_dir_cp' in df.columns:
    #     df['wind_dir'] = df['wind_dir_cp'].apply(lambda cp : cardinal_to_degrees(cp))
    #     df = df.drop(['wind_dir_cp'], axis='columns')

    # remove unused column names from the base dict
    return_cols = list(FIELDS_BASE.keys())
    return_cols.extend(
        [key for key in ['drought', 'ffdi','gfdi'] if key in col_names.values()]
    )
    return_cols = [key for key in return_cols if key in df.columns]
    
    return df[return_cols]

def gridded_to_df(fn: str, header: int = 6) -> DataFrame:
    """Reads BoM gridded weather forecast or observations into a pandas 
        `DataFrame`.
    
    Fields are mapped using

    > ```
    FIELDS_GRIDDED = {
        'date': 'Local Date',
        'time': 'Local Time',
        'temp': 'Temp (C)',
        'humidity': 'RH (%)',
        'wind_dir': 'Wind Dir',
        'wind_speed': 'Wind Speed (km/h)',
        'drought': 'Drought Factor',
        'ffdi': 'FFDI',
        'gfdi': 'GFDI',
    }```

    Args:
        fn: the path to the csv file

    Returns:
        a pandas DataFrame with the columns defined in `FIELDS_GRIDDED`
    """
    return weather_to_df(fn, header = header, col_names = FIELDS_GRIDDED, datetime_format="%d/%m/%Y %H:%M")

def fbcalcs_to_df(fn: str) -> DataFrame:
    """Reads weather data from FireBehaviourCalcs xlsm.

    Args:
        fn: the path to the FireBehaviourCalcs xlsm

    Returns:
        a pandas `DataFrame` suitable for use in `spreadmodels.py`
    """
    fields = {
        'date': 'B',
        'time': 'C',
        'temp': 'D',
        'humidity': 'E',
        'wind_speed': 'G',
        'wind_dir': 'F',
        'drought': 'H'
    }

    df = pd.DataFrame()

    if check_filepath(fn, suffix='xlsm'):
        wb = load_workbook(fn, data_only=True)
        ws = wb['Weather_Site']
        for field, col in fields.items():
            data = ws[col][11:]
            data = [cell.value for cell in data if isinstance(
                    cell.value, (float, int)
                )]
            df[field] = data
    
    return df

def df_to_weather(df: DataFrame, fn: str, col_names = FIELDS_BASE, datetime_format="%Y%m%d %H:%M", encoding=None) -> DataFrame:
    """"Creates a weather `DataFrame` containing the fields in `col_names`.

    Writes the weather data to a `*.csv` file
    
    The default fields are:
    ```
        'date': 'Local Date',
        'time': 'Local Time',
        'temp': 'Temp (C)',
        'humidity': 'RH (%)',
        'wind_dir': 'Wind Dir',
        'wind_speed': 'Wind Speed (km/h)',
        'drought': 'Drought Factor',
        'ffdi': 'FFDI',
        'gfdi': 'GFDI',
    ```
    
    If `datetime_format == 'iso8601'` the output format will be `%Y-%m-%dT%H:%M:%s+11:00`.

    Args:
        df: a `DataFrame` containing the weather and other fields
        fn: path to the output file
        col_names: the fields to be returned
        datetime_format: the output format for datetimes, dates and times.
        encoding: the coding for the output file if `*.csv`. Default is `UTF-8`

    Returns:
        a pandas `DataFrame` with the specified fields
    """
    df = df.copy()
    if 'date' in col_names.keys():
        df['date'] = df['date_time'].dt.strftime(datetime_format.split()[0])
    if 'time' in col_names.keys():
        df['time'] = df['date_time'].dt.strftime(datetime_format.split()[1])

    if 'date_time' in col_names.keys():
        df['date_time'] = df['date_time'].dt.strftime(datetime_format)
        # df['date_time'] = df['date_time'].dt.strftime("%d/%m/%Y %H:%M")

    
    # remove unused column names from the base dict
    col_names = {key:val for key,val in col_names.items() if key in df.columns}
    df = df[col_names.keys()]
    df = df.rename(columns=col_names)
    
    df.to_csv(fn, index=False, encoding=encoding)
    return df

def df_to_amicus(df, fn: str) -> DataFrame:
    """"Creates a weather `DataFrame` containing 
    [Amicus](https://research.csiro.au/amicus/) compatible fields.

    Writes the weather data to a `*.csv` file
    
    The output fields are:
    ```
    'date_time': 'Date time',
    'temp': 'Air temperature (°C)',
    'humidity': 'Relative humidity (%)',
    'wind_speed': '10 m wind speed (km/h)',
    'wind_dir': 'Wind direction (°)'
    ```
    
    Args:
        df: a `DataFrame` containing the weather and other fields
        fn: path to the output file

    Returns:
        a pandas `DataFrame` with Amicus compatible fields
    """

    return df_to_weather(df, fn, col_names=FIELDS_AMICUS, datetime_format="%d/%m/%Y %H:%M", encoding='cp1252')

def afdrs_meteograms_to_df(path) -> DataFrame:
    """"Creates a `DataFrame` from the indivudual csv downloads from
    AFDRS viewer metoegrams

    Writes the data to a `*.csv` file
    
    The output fields are:
    ```
    'date_time': 'Date time',
    'date': 'Date'
    'time': 'Time',
    'temp': 'Air temperature (°C)',
    'humidity': 'Relative humidity (%)',
    'wind_speed': '10 m wind speed (km/h)',
    'wind_dir': 'Wind direction (°)'
    ```
    
    Args:
    
    Returns:
        a pandas `DataFrame` with Amicus compatible fields
    """

    fields = {
        'date_time': 'Date time',
        'date': 'Date',
        'time': 'Time',
        'temp': 'Air temperature (°C)',
        'humidity': 'Relative humidity (%)',
        'wind_dir': 'Wind direction (°)',
        'wind_speed': '10 m wind speed (km/h)',
    }

    filenames = (
        'Relative Humidity.csv',
        'Wind Direction.csv',
        'Wind Speed.csv'
    )

    df = pd.read_csv(path / 'Temperature.csv')
    
    for fn in filenames:
        _df = pd.read_csv(path / fn)
        df = df.merge(_df, how='inner')

    return df


if __name__ == '__main__':
    path = Path(r"C:\Users\geoffg\Documents\Incidents\20221203_Redhead\AFDRS_downloads")
    df = afdrs_meteograms_to_df(path)
    print(df.head())

    out_path = Path()
