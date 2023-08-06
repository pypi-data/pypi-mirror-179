# -*- coding: utf-8 -*-
"""
Created on Mon Nov 14 20:10:48 2022

@author: MudithaRajapaksha
"""
import openpyxl
from openpyxl.worksheet.table import Table
import re
class ExcelTableMaster:
    
    
    def __init__(self, fileName , workSheetIndex , tableName):
        self.fileName = fileName
        self.workSheetIndex = workSheetIndex
        self.tableName = tableName
        
        file = open(self.fileName, "rb")
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.worksheets[self.workSheetIndex]
        return
    
    def writeTableRow(self, dataList):
        self.ws.append(dataList)
        self.ws.tables[self.tableName].ref = self.updatedRef(self.ws.tables[self.tableName].ref)
        self.wb.save(self.fileName)
        
    def writeMultipleTableRow(self, dataList):
        for row in dataList:
            self.ws.append(row)
        self.ws.tables[self.tableName].ref = self.updatedRef(self.ws.tables[self.tableName].ref)
        self.wb.save(self.fileName)    
        
    def writeUniqueTableRow(self ,keyColumnLetter, keyIndexNo ,dataList):
        valueFound = self.ValueExist(str(dataList[keyIndexNo]) , keyColumnLetter)
        if(valueFound == False):
            self.ws.append(dataList)
            
        self.ws.tables[self.tableName].ref = self.updatedRef(self.ws.tables[self.tableName].ref)
        self.wb.save(self.fileName)
            
    def writeMultipleUniqueTableRow(self,keyColumnLetter, keyIndexNo ,dataList):
        for row in dataList:
            valueFound = self.ValueExist(str( row[keyIndexNo] ), keyColumnLetter)
            if(valueFound == False):
                self.ws.append(row)
            
        self.ws.tables[self.tableName].ref = self.updatedRef(self.ws.tables[self.tableName].ref)
        self.wb.save(self.fileName)        
           
    
    def ValueExist(self ,checkValue , columnLetter):
        lastUsedRow = len(self.ws[columnLetter])
        startRow = int(re.search('[0-9]+', self.ws.tables[self.tableName].ref.split(":")[0]).group(0))
        
        valueFound = False
        for i in range(startRow , lastUsedRow + 1):
            rowVal = str(self.ws.cell(row=i, column=self.colNameToNum(columnLetter)).value)
            if( rowVal == checkValue):
                valueFound = True
                break 
        return valueFound
    
        
       
    #----------------------------------------- Support Methods--------------------------------------------------#    
    def updatedRef(self,oldRef):
        lastUsedRow = len(self.ws['A'])
        startLetters = re.search('[A-Z]+', oldRef.split(":")[0]).group(0)    
        endLetters = re.search('[A-Z]+', oldRef.split(":")[1]).group(0) 
        startRow = re.search('[0-9]+', oldRef.split(":")[0]).group(0)   
        newRef = "{}{}:{}{}".format(startLetters,startRow , endLetters,lastUsedRow)
        
        #update context ref
        return newRef

    def colNameToNum(self,name):
        pow = 1
        colNum = 0
        for letter in name[::-1]:
                colNum += (int(letter, 36) -9) * pow
                pow *= 26
        return colNum

