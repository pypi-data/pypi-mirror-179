from re import search
from pathlib import Path
from datetime import time, date

from openpyxl import load_workbook


class Data:
    # "_content_meta_data" contains information on what data is stored
    # where in the excel file and regex to extract it
    _c_md = {'User': {'loc': 'A3', 'reg': r'^User: (.*)$'},
             'Path': {'loc': 'A4', 'reg': r'^Path: (.*)$'},
             'Test ID': {'loc': 'A5', 'reg': r'^Test ID: (.*)$'},
             'Test Name': {'loc': 'A6', 'reg': r'^Test Name: (.*)$'},
             'Date': {'loc': 'A7', 'reg': r'^Date: (\d{1,2})/(\d{1,2})/(\d{4})$'},
             'Time': {'loc': 'A8', 'reg': r'^Time: (\d{1,2}):(\d{1,2}):(\d{1,2}) ([A-Z]{2})$'},
             'ID1': {'loc': 'A9', 'reg': r'^ID1: (.*)$'},
             'ID2': {'loc': 'A10', 'reg': r'^ID2: (.*)$'},
             'ID3': {'loc': 'A11', 'reg': r'^ID3: (.*)$'}}

    def __init__(self, path: Path) -> None:
        self.meta_data = []
        self.name = path.stem

        self.read_file(path)

    def read_file(self, path):
        wb = load_workbook(str(path))
        self.read_meta_data(wb.active)

    def read_meta_data(self, ws):
        for md, info in Data._c_md.items():
            raw = ws[info['loc']].value
            groups = search(info['reg'], raw).groups()

            if md == 'Time':
                hours = int(groups[0]) + 12*int(groups[3] == 'PM')
                minutes = int(groups[1])
                seconds = int(groups[2])
                value = time(hours, minutes, seconds)
            elif md == 'Date':
                day = int(groups[1])
                month = int(groups[0])
                year = int(groups[2])
                value = date(year, month, day)
            elif md == 'Path':
                value = Path(groups[0])
            else:
                value = groups[0]

            self.meta_data.append((md, value))
