import openpyxl
import os
import logging

logging.basicConfig(format='%(asctime)s [%(pathname)s:%(lineno)d] %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    level=logging.INFO)


def map_to_save_excel(results, len_urls, save_root):
    if not os.path.exists(save_root):
        os.makedirs(save_root)
    keys = list(results.keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '测试结果'
    row = 1
    column = 1
    for key in keys:
        ws.cell(row=row, column=column).value = key
        for sub in results[key]:
            sub_keys = list(sub.keys())
            sub_keys[0] = sub_keys[0] + ',' + str(len_urls)
            sub_values = list(sub.values())
            for k in sub_keys:
                sub_column = sub_keys.index(k)
                ws.cell(row=row + 1, column=sub_column + 1).value = k
            for v in sub_values:
                sub_column = sub_values.index(v)
                ws.cell(row=row + 2, column=sub_column + 1).value = v
            row += 2
        row += 1
    wb.save(os.path.join(save_root, 'results.xlsx'))
    logging.info('{}/results.xlsx save done.'.format(save_root))
