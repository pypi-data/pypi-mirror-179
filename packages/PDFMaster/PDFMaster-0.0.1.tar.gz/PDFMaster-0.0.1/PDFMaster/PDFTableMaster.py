# -*- coding: utf-8 -*-
"""
Created on Tue Nov 15 12:27:26 2022

@author: MudithaRajapaksha
"""


import glob
import re
from datetime import datetime
import PyPDF2
from PyPDF2 import PdfFileReader
from openpyxl import load_workbook
from matplotlib import patches
import matplotlib.pyplot as plt
import pdfquery
import pdfminer
from pdfminer.pdfpage import PDFPage, PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import PDFPageAggregator
import os
from pathlib import Path
from PyPDF2 import PdfFileReader, PdfFileWriter
import pandas as pd
import sys
import warnings
from win32com.client import Dispatch
from abc import ABC, abstractmethod  


warnings.filterwarnings("ignore")

class PDFTableMaster:

    #Default Parameters
    upperBoundry = 5
    lowerBoundry =5
    margin = 0.3
   
    
    def __init__(self, fileName):
        self.fileName = fileName
        
        self.TEXT_ELEMENTS = [
            pdfminer.layout.LTTextBox,
            pdfminer.layout.LTTextBoxHorizontal,
            pdfminer.layout.LTTextLine,
            pdfminer.layout.LTTextLineHorizontal
        ]
        
        
        return
    
    
    def extract_PDF_Pages_To_NewFile(self, pages ):
        file = self.fileName
        segments = [{'start': pages[0], 'end': pages[-1]}]
        file_name = os.path.basename(file)
    
        with open(file, 'rb') as read_stream:
            pdf_reader = PdfFileReader(read_stream)
            for segment in segments:
                pdf_writer = PdfFileWriter()
                # support {'start': 3, 'end': 3} or (start, end)
                try:
                    start_page, end_page = segment['start'], segment['end']
                except TypeError:
                    start_page, end_page = segment
                for page_num in range(start_page - 1, end_page):
                    pdf_writer.addPage(pdf_reader.getPage(page_num))
                p = Path(file)
                #ouput = p.parent / p.with_stem(f'{p.stem}_pages_{start_page}-{end_page}')
                ouput = "temp/{}".format(file_name)
                with open(ouput, 'wb') as out:
                    pdf_writer.write(out)
        #New File Name                        
        return ouput
    
    def extract_page_layouts(self):
        """
        Extracts LTPage objects from a pdf file.
        modified from: http://www.degeneratestate.org/posts/2016/Jun/15/extracting-tabular-data-from-pdfs/
        Tests show that using PDFQuery to extract the document is ~ 5 times faster than pdfminer.
        """
        laparams = LAParams()
    
        with open(self.fileName, mode='rb') as pdf_file:
            print("Open document %s" % pdf_file.name)
            document = pdfquery.PDFQuery(pdf_file).doc
    
            if not document.is_extractable:
                raise PDFTextExtractionNotAllowed
    
            rsrcmgr = PDFResourceManager()
            device = PDFPageAggregator(rsrcmgr, laparams=laparams)
            interpreter = PDFPageInterpreter(rsrcmgr, device)
    
            layouts = []
            for page in PDFPage.create_pages(document):
                interpreter.process_page(page)
                layouts.append(device.get_result())
        
        self.layouts = layouts
        return layouts
    
    
    
    def get_data_rows_list(self , pages = [0]  ,clean =  True , cleanMaster = None):
        page_layouts = self.extract_page_layouts()
        pageFinalList = []
        for page in pages:
            
            current_page = page_layouts[page]
            
            text_objects =  PDFTableMaster.get_text_objects(current_page)
            rect_objects =  PDFTableMaster.get_rect_objects(current_page)
            characters = PDFTableMaster.extract_characters(text_objects , self.TEXT_ELEMENTS)
            #PDFTableMaster.plotPDF(current_page,text_objects,rect_objects ,characters)
            sorted_rows = PDFTableMaster.arrange_text(characters)
            row_texts = PDFTableMaster.extract_text(sorted_rows)
            text = PDFTableMaster.arrange_and_extract_text(characters)
           
            if(clean == False):
                cleaned_List = text
            else:
                #cleaned_List = PDFTableMaster.clean_lists(text)
                cleaned_List = cleanMaster.cleanListMaster(text)
                
            pageFinalList.extend(cleaned_List)
    
        return pageFinalList
    
    
    def get_data_rows_pandas(self , pages = [0] , columns = ['A', 'B'] , clean =  True , cleanMaster = None):
        table_rows_list = self.get_data_rows_list(pages , clean , cleanMaster)
        
        df = pd.DataFrame(columns=columns)
        for row in table_rows_list:
            temp = pd.DataFrame([row], columns=columns)
            df = df.append(temp, ignore_index=True)
        return df
    
    
    
    def write_Data_To_Excel(self , pages = [0] , saveFilePath = "data.xlsx"):
        file = self.fileName
        finalPd = self.get_data_rows_pandas(pages)
        
        print("=======WRITEING TO EXCEL========")
       
        #userName = getpass.getuser()
        #filePath = "C:/Users/{}/Lynear/LYNEAR Research - Documents/Statistics/Sri Lanka/Product Prices/".format(userName)
        #saveFilePath =  "data.xlsx"
        
        sheetName = PDFTableMaster.extract_information(self.fileName)
    
        wb = load_workbook(saveFilePath, read_only=True)  # open an Excel file and return a workbook
        if sheetName in wb.sheetnames:
            pass
        else:
            # finalPd.to_excel(fullPath, sheet_name=sheet_name)
            with pd.ExcelWriter(saveFilePath, engine="openpyxl", mode='a') as writer:
                finalPd.to_excel(writer, sheet_name=sheetName , index=False)
    
    
    def set_parameters(self , parameters):
        PDFTableMaster.upperBoundry = parameters["upperBoundry"]
        PDFTableMaster.lowerBoundry = parameters["lowerBoundry"]
        PDFTableMaster.margin = parameters["margin"]
        
    
    def table_exists(self , pageNo = 1):
        import tabula
        try:
            table = tabula.read_pdf(self.fileName,pages=pageNo)
        except:
            return False
        
        if(len(table) > 0):
            return True
        else:
            return False
        
            
    def citeria_exist(self ,citerias = ["TEST" , "test" , "T" , "t"] ):
        # Locate the page rate to be copied
        object = PyPDF2.PdfFileReader(self.fileName)  # open the pdf file
        NumPages = object.getNumPages()  # get number of pages
        pages = []
    
        for i in range(NumPages):  # extract text and do the search
            PageObj = object.getPage(i)
            #print("this is page " + str(i + 1) , end = " - ")
            pageText = PageObj.extractText()
            # print(Text)

            satisfied = []
            for citeria in citerias:
                if (re.search(citeria, pageText)):
                    satisfied.append(True)
                else:
                    satisfied.append(False)

            if (False not in satisfied):
                pages.append(i+1)
                #print("Page Identified")
            else:
                pass
                #print('Not Found')

        print("Trade Pages  ", pages)
        if(len(pages)>0):
            return True
        else:
            return False
        
    
    def identify_Pages_To_Scrape(self  , citerias = ["TEST" , "test" , "T" , "t"]):
            # Locate the page rate to be copied
            object = PyPDF2.PdfFileReader(self.fileName)  # open the pdf file
            NumPages = object.getNumPages()  # get number of pages
            pages = []
        
            for i in range(0, NumPages-1):  # extract text and do the search
                PageObj = object.getPage(i)
                #print("this is page " + str(i + 1) , end = " - ")
                pageText = PageObj.extractText()
                # print(Text)
    
                satisfied = []
                for citeria in citerias:
                    if (re.search(citeria, pageText)):
                        satisfied.append(True)
                    else:
                        satisfied.append(False)
    
                if (False not in satisfied):
                    pages.append(i+1)
                    #print("Page Identified")
                else:
                    pass
                    #print('Not Found')
    
            print("Trade Pages  ", pages)
            return pages
    
    
    #-------------------------------------------------------Static Methods ---------------------------------------------------------
    
    
    
    
    @staticmethod
    def get_text_objects(page_layout):
        texts = []
        # seperate text and rectangle elements
        for elem in page_layout:
            if isinstance(elem, pdfminer.layout.LTTextBoxHorizontal):
                texts.append(elem)
        return texts
    
    @staticmethod
    def get_rect_objects(page_layout):
        rect = []
        # seperate text and rectangle elements
        for elem in page_layout:
            if isinstance(elem, pdfminer.layout.LTRect):
                rect.append(elem)
        return rect
    
    @staticmethod
    def flatten(lst):
        """Flattens a list of lists"""
        return [item for sublist in lst for item in sublist]
    
    
    @staticmethod
    def extract_characters(element , TEXT_ELEMENTS):
        """
        Recursively extracts individual characters from
        text elements.
        """
        if isinstance(element, pdfminer.layout.LTChar):
            return [element]
    
        if any(isinstance(element, i) for i in TEXT_ELEMENTS):
            return PDFTableMaster.flatten([PDFTableMaster.extract_characters(e , TEXT_ELEMENTS) for e in element])
    
        if isinstance(element, list):
            return PDFTableMaster.flatten([PDFTableMaster.extract_characters(l , TEXT_ELEMENTS) for l in element])
    
        return []
    
    @staticmethod
    def arrange_and_extract_text(characters):
        margin= PDFTableMaster.margin
        
        rows = sorted(list(set(c.bbox[1] for c in characters)), reverse=True)
    
        row_texts = []
        for row in rows:
            #sorted_row = sorted([c for c in characters if c.bbox[1] == row], key=lambda c: c.bbox[0])
            sorted_row = sorted([c for c in characters if PDFTableMaster.proximityRows(c.bbox[1], row)], key=lambda c: c.bbox[0])
    
            col_idx = 0
            row_text = []
            for idx, char in enumerate(sorted_row[:-1]):
                if (sorted_row[idx + 1].bbox[0] - char.bbox[2]) > margin:
                    col_text = "".join([c.get_text() for c in sorted_row[col_idx:idx + 1]])
                    col_idx = idx + 1
                    row_text.append(col_text)
                elif idx == len(sorted_row) - 2:
                    col_text = "".join([c.get_text() for c in sorted_row[col_idx:]])
                    row_text.append(col_text)
            row_texts.append(row_text)
        return row_texts
    
    @staticmethod
    def arrange_text(characters):
        """
        For each row find the characters in the row
        and sort them horizontally.
        """
    
        # find unique y0 (rows) for character assignment
        rows = sorted(list(set(c.bbox[1] for c in characters)), reverse=True)
        #print(rows)
        for c in characters:
            x= c.bbox[1]
            #print(x)
            #print(type(x))
    
        sorted_rows = []
        for row in rows:
            #sorted_row = sorted([c for c in characters if c.bbox[1] == row], key=lambda c: c.bbox[0])
            sorted_row = sorted([c for c in characters if PDFTableMaster.proximityRows(c.bbox[1],row ) ], key=lambda c: c.bbox[0])
            sorted_rows.append(sorted_row)
    
        return sorted_rows
    
    @staticmethod
    def proximityRows(r1 , r2):
        lower = r1- PDFTableMaster.upperBoundry
        upper = r1 + PDFTableMaster.lowerBoundry
    
        if(r2 > lower and r2 < upper):
            return True
        else:
            return False
    
    @staticmethod
    def extract_text(rows):
        row_texts = []
        for row in rows:
            row_text = ["".join([c.get_text() for c in row])]
            row_texts.append(row_text)
        return row_texts
    
   
    @staticmethod
    def build_dict_hierarchy(texts):
        data = {
            'General Info': {
                'Name': texts[0][0],
                'Company Name': texts[1][0],
                'Issue': texts[2][0][:-5],
                'Year': texts[2][0][-4:]
            }
        }
    
        LABELS = [
            'Heading 1 is short',
            'Heading 2 which is actually very long, much longer than the first heading',
            'Heading 3 is also a little bit longer but not as long as the second'
        ]
    
        SUBLABELS = [
            'Segment 1-1',
            'Segment 1-2',
            'Segment 1-3',
            'Segment 2-1'
        ]
    
        SUMS = [
            'Something:',
            'Something else:',
            'Sum:'
        ]
    
        key = None
        subkey = None
        for row in texts[4:]:
            # assure that row isn't empty
            if row:
                word = row[0]
            else:
                continue
    
            if word in LABELS:
                key = word
                subkey = None
                data[key] = {}
            elif word in SUBLABELS:
                subkey = word
                data[key][subkey] = {}
            elif word in SUMS:
                data[key]['Sum'] = row[-1]
            elif subkey:
                ID = " ".join(row[:-1])
                data[key][subkey][ID] = row[-1]
            elif key:
                ID = " ".join(row[:-1])
                data[key][ID] = row[-1]
    
        return data
    
    @staticmethod
    def draw_rect_bbox(bbox, ax, color):
        """
        Draws an unfilled rectable onto ax.
        """
        x0, y0, x1, y1 = tuple(bbox)
        ax.add_patch(
            patches.Rectangle(
                (x0, y0),
                x1 - x0,
                y1 - y0,
                fill=False,
                color=color
            )
        )
    
    @staticmethod
    def draw_rect(rect, ax, color="black"):
        PDFTableMaster.draw_rect_bbox(rect.bbox, ax, color)
    
    
    @staticmethod
    def plotPDF(current_page ,text_objects, rect_objects , characters):
        xmin, ymin, xmax, ymax = current_page.bbox
        size = 6
        num_pages = 2
    
        fig, axes = plt.subplots(1, num_pages, figsize=(num_pages * size, size * (ymax / xmax)), sharey=True, sharex=True)
    
        # rects and chars
        ax = axes[0]
        for rect in rect_objects:
            PDFTableMaster.draw_rect(rect, ax)
    
        for c in characters:
            PDFTableMaster.draw_rect(c, ax, "red")
    
        # chars and TextBoxes
        ax = axes[1]
        for c in characters:
            PDFTableMaster.draw_rect(c, ax, "red")
    
        for textbox in text_objects:
            PDFTableMaster.draw_rect(textbox, ax, "blue")
    
        plt.xlim(xmin, xmax)
        plt.ylim(ymin, ymax)
        #plt.show()
        plt.savefig('temp.png')
        
        

        
    @staticmethod    
    def getPDFList(folderPath):
            '''
            data_dir = 'data/'
    
            if len(sys.argv) == 2:
                file = data_dir + sys.argv[1]
                print(f"using file {file}")
            else:
                file = data_dir + "test2.pdf"
            '''
            # userName = getpass.getuser()
            # filePath = "C:/Users/{}/Lynear/LYNEAR Research - Documents/Statistics/Sri Lanka/Product Prices/".format(userName)
    
            #fileslist = glob.glob(filePath+"*.pdf")
            fileslist = glob.glob("{}/*.pdf".format(folderPath))
            fileNameList = []
            for file in fileslist:
                fileNameList.append(file.split("\\")[-1])
            
            fileslist.sort()
            fileNameList.sort()
            
            return [fileslist , fileNameList]
        
    @staticmethod
    def extract_information(pdf_path):
        with open(pdf_path, 'rb') as f:
            pdf = PdfFileReader(f)
            information = pdf.getDocumentInfo()
            number_of_pages = pdf.getNumPages()
    
        txt = f"""
        Information about {pdf_path}: 
    
        Author: {information.author}
        Creator: {information.creator}
        Producer: {information.producer}
        Subject: {information.subject}
        Title: {information.title}
        Number of pages: {number_of_pages}
        """
    
        print(txt)
        return information.title
    

    @staticmethod
    def close_already_open_excel():
        import os
    
        def closeFile():
    
            try:
                os.system('TASKKILL /F /IM excel.exe')
    
            except Exception:
                print("KU")
    
        closeFile()    
        
    

    #----------------------------------------------- Abstract Classes --------------------------------------------------------
class CleanMaster(ABC):
        
        
        @staticmethod
        @abstractmethod
        def cleanListMaster(self , rows):
            pass
        
        
        
        # Use following commented methods as draft to implement the class
        #shoud accept list of lists and return a lists of lists
        
        
        '''
        def clean_lists(rows):
            def digitCount(row):
                count = 0
                for item in row:
                    if (bool(re.search(r'\d', item))):
                        count = count + 1
                        # print("yes - " , item)
                    else:
                        pass
                        # print("No - " , item)
                return count
        
            companyRowList = []
            uniqueComPanyList = []
            for row in rows:
                if(len(row) >= PDFTableMaster.numberOfColumns[0] and len(row) <= PDFTableMaster.numberOfColumns[1]):
                    if(row[0] not in uniqueComPanyList):
                        cleanedRow = []
                        # Irregular Items Combined into one
                        result = PDFTableMaster.irregularRows(row[6])
                        if(result != None):
                            row[6] = result[0]
                            row[7] = result[1]
        
                        result = PDFTableMaster.irregularRows(row[7])
                        if (result != None):
                            row[7] = result[0]
                            row[8] = result[1]
        
                        for item in range(0 , len(row) -1):
                            itemC = row[item].strip().replace("," , "")
                            if(item != 0):
                                itemC = re.sub(r"\s+", "", itemC, flags=re.UNICODE)
                            cleanedRow.append(itemC)
        
                        if(digitCount(cleanedRow) >= 5 and cleanedRow[0] not in uniqueComPanyList):
                            companyRowList.append(cleanedRow)
                            uniqueComPanyList.append(cleanedRow[0])
        
        
            finalPageList = []
            for row in companyRowList:
                mainCols = [row[0]  , row[1] , row[6] , row[7]]
                finalPageList.append(mainCols)
                print(mainCols)
        
            return finalPageList
        
        def irregularRows(strIrregularNumber):
            strIrregularNumber = strIrregularNumber.replace(" ","")
            def deriveTwoNumbers(strNumberBind, re):
                numCount = re.end() - re.start() - 2  # commas -2
                num1 = strNumberBind[:re.start() + 4]
                num2 = strNumberBind[re.start() + 4:]
                return [num1, num2]
        
            re1 = re.search(r'\,\d\d\d\d\d\d\,', strIrregularNumber)
            re2 = re.search(r'\,\d\d\d\d\d\,', strIrregularNumber)
            re3 = re.search(r'\,\d\d\d\d\,', strIrregularNumber)
        
            if (re1 != None):
                return deriveTwoNumbers(strIrregularNumber, re1)
            elif (re2 != None):
                return deriveTwoNumbers(strIrregularNumber, re2)
            elif (re3 != None):
                return deriveTwoNumbers(strIrregularNumber, re3)
            else:
                return None
            
            '''
    
#----------------------------------------------------------------------------------------------------------------------------------------------------




    
    
    
    
